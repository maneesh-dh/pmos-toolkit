"""ingest.py — CSV / TSV / XLSX / XLS / PDF readers -> normalised rows.

Dependencies:
- stdlib only for CSV / TSV.
- `openpyxl >= 3.0` for .xlsx / .xls (auto-detect missing; surface install
  instructions).
- PDF reading is best-effort and falls through to a Read-tool LLM
  tabulation pass (handled by the parent skill, not this module).

Functions
---------
read_responses(path, sheet=None) -> (rows, header_order)
    Dispatch by extension. Returns a list of dicts (one per response row)
    and a list of column names in input order.

CLI: python3 -m helpers.ingest --selftest
"""
from __future__ import annotations
import csv
import os
import sys


def _read_csv(path: str, delimiter: str | None = None) -> tuple[list[dict], list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(2048)
        f.seek(0)
        if delimiter is None:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ","
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = [{k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()}
                for r in reader]
        headers = list(reader.fieldnames or [])
    return rows, headers


def _read_xlsx(path: str, sheet) -> tuple[list[dict], list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "survey-analyse: openpyxl not installed. Run: "
            "python3 -m pip install 'openpyxl>=3.0'"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    headers_row = next(it, ())
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(headers_row)]
    rows: list[dict] = []
    for row in it:
        d = {}
        for h, v in zip(headers, row):
            if isinstance(v, str):
                v = v.strip()
            d[h] = "" if v is None else v
        # Skip fully empty rows.
        if any(v not in ("", None) for v in d.values()):
            rows.append(d)
    return rows, headers


def read_responses(path: str, sheet=None) -> tuple[list[dict], list[str]]:
    """Dispatch by extension. Returns (rows, header_order)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv",):
        return _read_csv(path)
    if ext in (".tsv",):
        return _read_csv(path, delimiter="\t")
    if ext in (".xlsx", ".xlsm", ".xls"):
        return _read_xlsx(path, sheet)
    if ext == ".pdf":
        raise NotImplementedError(
            "PDF ingestion is best-effort and goes through the Read-tool "
            "fallback in the parent skill — call read_responses() with a "
            ".csv/.tsv/.xlsx instead. See SKILL.md Phase 1 for the PDF path."
        )
    raise ValueError(f"Unsupported response file extension: {ext}")


def _selftest() -> int:
    import tempfile
    # CSV round-trip
    with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False) as f:
        f.write("plan,nps\nFree,10\nPro,8\n\n")
        p = f.name
    try:
        rows, headers = read_responses(p)
        assert headers == ["plan", "nps"], headers
        assert len(rows) == 2 and rows[0]["plan"] == "Free", rows
    finally:
        os.unlink(p)
    # TSV
    with tempfile.NamedTemporaryFile("w", suffix=".tsv", delete=False) as f:
        f.write("a\tb\n1\t2\n")
        p = f.name
    try:
        rows, headers = read_responses(p)
        assert rows == [{"a": "1", "b": "2"}], rows
    finally:
        os.unlink(p)
    # PDF: NotImplementedError surfaces, not silently empty
    try:
        read_responses("dummy.pdf")
    except NotImplementedError:
        pass
    else:
        raise AssertionError("expected NotImplementedError for PDF")
    print("ingest.read_responses: OK (CSV/TSV; PDF deferred to parent)")
    return 0


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        sys.exit(_selftest())
    print("usage: python3 -m helpers.ingest --selftest")
    sys.exit(64)
